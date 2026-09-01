from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

from schedule.models import StudentGroups


@dataclass
class GroupIndex:
    """Maps students and named groups from the UFAR Excel roster."""

    students: dict[str, StudentGroups] = field(default_factory=dict)
    group_aliases: dict[str, set[str]] = field(default_factory=dict)  # alias -> column keys
    known_labels: list[str] = field(default_factory=list)


def _norm(s: str) -> str:
    s = (s or "").casefold().strip()
    s = s.replace("ё", "е")
    s = re.sub(r"[^\w\s]+", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s)
    return s


def _is_person_name(value: object) -> bool:
    if not isinstance(value, str):
        return False
    text = value.strip()
    if not text or text.isdigit():
        return False
    if any(ch.isdigit() for ch in text):
        return False
    # Heuristic: at least two tokens (surname + name)
    parts = text.split()
    return len(parts) >= 2 and all(any(c.isalpha() for c in p) for p in parts)


def find_fallback_xlsx(fallback_dir: Path) -> Optional[Path]:
    files = sorted(fallback_dir.glob("*.xlsx"))
    return files[0] if files else None


def load_group_index(xlsx_path: Path) -> GroupIndex:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    index = GroupIndex()
    students: dict[str, StudentGroups] = {}

    def ensure(name: str) -> StudentGroups:
        key = _norm(name)
        if key not in students:
            students[key] = StudentGroups(display_name=name.strip())
        return students[key]

    # --- TP խմբեր ---
    tp_sheet = next((n for n in wb.sheetnames if n.strip().upper().startswith("TP")), None)
    if tp_sheet:
        ws = wb[tp_sheet]
        # Detect "Խումբ N" / "Groupe N" headers anywhere; names sit below in nearby columns.
        headers: list[tuple[int, int, str]] = []  # row, col, TP_n
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=min(ws.max_column, 12)):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                m = re.search(r"(խումբ|groupe|group)\s*(\d+)", val, re.I)
                if m:
                    headers.append((cell.row, cell.column, f"TP_{int(m.group(2))}"))
        for hrow, hcol, tp_key in headers:
            name_col = hcol + 1 if hcol + 1 <= ws.max_column else hcol
            for r in range(hrow + 1, min(hrow + 30, ws.max_row + 1)):
                # stop at next header in this column
                nxt = ws.cell(r, hcol).value
                if isinstance(nxt, str) and re.search(r"խումբ|groupe|group", nxt, re.I):
                    break
                left = ws.cell(r, name_col).value
                if not _is_person_name(left) and hcol >= 1:
                    left = ws.cell(r, hcol).value
                if _is_person_name(left):
                    ensure(str(left)).tp = tp_key
        # Fallback to the known 2026 layout if headers were not found
        if not headers:
            for row in range(5, 28):
                left = ws.cell(row, 2).value
                right = ws.cell(row, 6).value
                if _is_person_name(left):
                    ensure(str(left)).tp = "TP_1"
                if _is_person_name(right):
                    ensure(str(right)).tp = "TP_3"
            for row in range(30, 54):
                left = ws.cell(row, 2).value
                right = ws.cell(row, 6).value
                if _is_person_name(left):
                    ensure(str(left)).tp = "TP_2"
                if _is_person_name(right):
                    ensure(str(right)).tp = "TP_4"

    # --- TD խմբեր ---
    td_sheet = next((n for n in wb.sheetnames if n.strip().upper().startswith("TD")), None)
    if td_sheet:
        ws = wb[td_sheet]
        headers_td: list[tuple[int, int, str]] = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), max_col=min(ws.max_column, 12)):
            for cell in row:
                val = cell.value
                if not isinstance(val, str):
                    continue
                if re.search(r"avanc|advanc", val, re.I):
                    headers_td.append((cell.row, cell.column, "TD_ADV"))
                    continue
                m = re.search(r"(խումբ|groupe|group)\s*(\d+)", val, re.I)
                if m:
                    headers_td.append((cell.row, cell.column, f"TD_{int(m.group(2))}"))
        for hrow, hcol, td_key in headers_td:
            name_col = hcol + 1 if hcol + 1 <= ws.max_column else hcol
            for r in range(hrow + 1, min(hrow + 50, ws.max_row + 1)):
                nxt = ws.cell(r, hcol).value
                if isinstance(nxt, str) and re.search(r"խումբ|groupe|group|avanc", nxt, re.I):
                    break
                name = ws.cell(r, name_col).value
                if not _is_person_name(name):
                    name = ws.cell(r, hcol).value
                if _is_person_name(name):
                    ensure(str(name)).td = td_key
        if not headers_td:
            for row in range(5, 46):
                g1 = ws.cell(row, 2).value
                g2 = ws.cell(row, 5).value
                adv = ws.cell(row, 8).value
                if _is_person_name(g1):
                    ensure(str(g1)).td = "TD_1"
                if _is_person_name(g2):
                    ensure(str(g2)).td = "TD_2"
                if _is_person_name(adv):
                    ensure(str(adv)).td = "TD_ADV"

    # --- French IMA ---
    if "French IMA" in wb.sheetnames:
        ws = wb["French IMA"]
        current_room: Optional[str] = None
        current_group: Optional[str] = None
        for row in range(1, ws.max_row + 1):
            salle = ws.cell(row, 1).value
            name = ws.cell(row, 3).value
            notes = ws.cell(row, 5).value
            if isinstance(salle, (int, float)) or (isinstance(salle, str) and salle.strip().isdigit()):
                current_room = f"FR_{int(salle)}"
            if isinstance(notes, str) and "խումբ" in notes.casefold():
                m = re.search(r"(\d+)", notes)
                if m:
                    current_group = m.group(1)
            if _is_person_name(name) and current_room:
                sg = ensure(str(name))
                sg.french_room = current_room
                sg.french_group = current_group

    # Everyone is CM A for year 1 stream in this PDF
    for sg in students.values():
        sg.cm = "CM_A"

    index.students = students

    # Named group aliases
    aliases: dict[str, set[str]] = {}
    for n in range(1, 5):
        aliases[_norm(f"tp {n}")] = {f"TP_{n}", "CM_A"}
        aliases[_norm(f"tp{n}")] = {f"TP_{n}", "CM_A"}
        aliases[_norm(f"խումբ {n}")] = {f"TP_{n}", "CM_A"}
        aliases[_norm(f"խումբ{n}")] = {f"TP_{n}", "CM_A"}
    for n in range(1, 4):
        aliases[_norm(f"td {n}")] = {f"TD_{n}", "CM_A"}
        aliases[_norm(f"td{n}")] = {f"TD_{n}", "CM_A"}
    aliases[_norm("td advanced")] = {"TD_ADV", "CM_A"}
    aliases[_norm("td avancé")] = {"TD_ADV", "CM_A"}
    aliases[_norm("avancé")] = {"TD_ADV", "CM_A"}
    aliases[_norm("advanced")] = {"TD_ADV", "CM_A"}
    aliases[_norm("cm a")] = {"CM_A"}
    aliases[_norm("cm")] = {"CM_A"}
    for room in (101, 102, 103, 104, 200):
        aliases[_norm(f"french {room}")] = {f"FR_{room}", "CM_A"}
        aliases[_norm(f"fr {room}")] = {f"FR_{room}", "CM_A"}
        aliases[_norm(f"իապի {room}")] = {f"FR_{room}", "CM_A"}

    index.group_aliases = aliases
    index.known_labels = sorted(
        {
            *[sg.display_name for sg in students.values()],
            "TP 1",
            "TP 2",
            "TP 3",
            "TP 4",
            "TD 1",
            "TD 2",
            "TD 3",
            "TD Advanced",
            "CM A",
        }
    )
    return index


def resolve_group(index: GroupIndex, query: str) -> Optional[StudentGroups]:
    q = _norm(query)
    if not q:
        return None

    # Exact student name
    if q in index.students:
        sg = index.students[q]
        sg.raw_query = query
        return sg

    # Fuzzy student: all tokens of query appear in name, or vice versa
    best: Optional[StudentGroups] = None
    best_score = 0
    q_tokens = set(q.split())
    for key, sg in index.students.items():
        name_tokens = set(key.split())
        if not q_tokens or not name_tokens:
            continue
        overlap = len(q_tokens & name_tokens)
        if overlap == 0:
            continue
        score = overlap / max(len(q_tokens), len(name_tokens))
        if q in key or key in q:
            score += 0.5
        if score > best_score:
            best_score = score
            best = sg
    if best and best_score >= 0.6:
        best.raw_query = query
        return best

    # Named group alias
    if q in index.group_aliases:
        keys = index.group_aliases[q]
        tp = next((k for k in keys if k.startswith("TP_")), None)
        td = next((k for k in keys if k.startswith("TD_")), None)
        fr = next((k for k in keys if k.startswith("FR_")), None)
        return StudentGroups(
            display_name=query.strip(),
            cm="CM_A",
            tp=tp,
            td=td,
            french_room=fr,
            raw_query=query,
        )

    return None
